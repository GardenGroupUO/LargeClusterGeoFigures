import os, shutil
import numpy as np
import matplotlib.colors as mcolors

from ase.io import read, write
from ase.visualize import view

#from ase.data import atomic_numbers, chemical_symbols
#from asap3.analysis.rdf import RadialDistributionFunction

from collections import Counter

import matplotlib.pyplot as plt
from matplotlib.pyplot import figure

#import networkx as nx

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import itertools

try:
	from tqdm import tqdm, trange
	does_tqdm_exist = True
except ImportError:
	print('######################################################')
	print('######################################################')
	print('NOTE: You should think about installing tqdm for a sweet progress bar!')
	print('pip install --user --upgrade tqdm')
	print('######################################################')
	print('######################################################')
	does_tqdm_exist = False

from ase.io import write


def get_distance(vector1,vector2):
	"""
	Get the distance betweeen two vectors
	"""
	dist_xx = vector1[0] - vector2[0]
	dist_yy = vector1[1] - vector2[1]
	dist_zz = vector1[2] - vector2[2]
	distance = (dist_xx**2.0 + dist_yy**2.0 + dist_zz**2.0) ** 0.5
	return distance

class LargeClusterGeoFigures_Program:
	def __init__(self, r_cut, elements=['Cu','Pd'],focus_plot_with_respect_to_element='Cu',path_to_files='.',record_all_files=False,add_legend=False,bulk_colour='FFC0CB',face_colour='FF0000',vertex_colour='90EE90',edge_colour='ADD8E6',none_colour='FFFFFF',auto_centre=False,record_bond_distance=False,testing=False):
		"""
		This program is designed to analyse the nearest neighbours within a cluster based on specifications from Dr. Caitlin Casey-Stevens.

		More information about this program and the input parameters are given in the README.md file in Githab.
		"""
		# General information about running the program
		self.path_to_here = os.path.abspath(path_to_files)
		self.r_cut = r_cut
		self.elements = elements
		self.focus_plot_with_respect_to_element = focus_plot_with_respect_to_element
		self.record_all_files = record_all_files
		self.add_legend = add_legend
		self.record_bond_distance = record_bond_distance

		# The colours for all the different types of structural environments. 
		def rgb_to_hex(rgb):
			r, g, b = rgb
			r, g, b = int(r*255.0), int(g*255.0), int(b*255.0)
			return '#%02x%02x%02x' % (r, g, b)
		self.bulk_colour = (bulk_colour if isinstance(bulk_colour,str) else rgb_to_hex(bulk_colour)).upper() 
		self.face_colour = (face_colour if isinstance(face_colour,str) else rgb_to_hex(face_colour)).upper() 
		self.vertex_colour = (vertex_colour if isinstance(vertex_colour,str) else rgb_to_hex(vertex_colour)).upper() 
		self.edge_colour = (edge_colour if isinstance(edge_colour,str) else rgb_to_hex(edge_colour)).upper() 
		self.none_colour = (none_colour if isinstance(none_colour,str) else rgb_to_hex(none_colour)).upper() 
		self.colours = {'bulk': self.bulk_colour, 'face': self.face_colour, 'edge': self.edge_colour, 'vertex': self.vertex_colour, 'None': self.none_colour}

		# Determine if you want to autocentre your clusters in their analysed xyz files.
		self.auto_centre = auto_centre

		# Other general information the user does not need to provide.
		self.original_path = os.getcwd()
		self.types_of_NNs = ['bulk', 'face', 'edge', 'vertex']

		# information for debugging
		self.testing_name = 'TESTING_' if testing else ''

		self.run()

	def run(self):	
		################## remove folder ##################	
		workbook_name = self.testing_name+"LargeClusterGeo_Data_Path"+self.path_to_here.replace(self.original_path,'').replace('/','_')+'_focus_element_'+str(self.focus_plot_with_respect_to_element)
		workbook_cluster_folder = workbook_name+'_clusters'
		if os.path.exists(workbook_cluster_folder):
			shutil.rmtree(workbook_cluster_folder)
		workbook_cluster_traj_folder = workbook_name+'_traj_file_data'
		if os.path.exists(workbook_cluster_traj_folder):
			shutil.rmtree(workbook_cluster_traj_folder)
		###################################################
		print('--------------------------------------')
		print('Processing: '+str(self.path_to_here))
		print('--------------------------------------')
		print('Getting data from XYZ files')
		clusters_to_process, traj_files_to_process = self.get_paths_of_all_clusters_to_process()
		print('Finished getting data from XYZ files')
		print('--------------------------------------')
		print('Processing data')
		cluster_information = self.process_cluster_data(clusters_to_process)
		print('Processed data')
		print('--------------------------------------')
		print('Analysing data')
		analysed_cluster_information = self.analyse_cluster_data(cluster_information)
		print('Finished analysing data')
		print('--------------------------------------')
		print('Making excel spreadsheet')
		################### make folder ###################
		workbook_cluster_folder = workbook_name+'_clusters'
		if os.path.exists(workbook_cluster_folder):
			shutil.rmtree(workbook_cluster_folder)
		os.makedirs(workbook_cluster_folder)
		###################################################
		self.record_to_excel(cluster_information,analysed_cluster_information,workbook_name,combine_clusters_together=False,record_bond_distance=self.record_bond_distance,workbook_cluster_folder=workbook_cluster_folder,workbook_cluster_traj_folder=workbook_cluster_traj_folder)
		print('Finished excel spreadsheet')
		print('--------------------------------------')
		if len(traj_files_to_process) > 0:
			################### make folder ###################
			os.makedirs(workbook_cluster_traj_folder)
			###################################################
			print('--------------------------------------')
			print('Traj file exist. We assume these are trajectories of the same cluster during a MD simulation')
			print('Will now process each of these traj file in the same way as the xyz and vasp files.')
			self.process_traj_files(traj_files_to_process, cluster_information, workbook_cluster_folder=workbook_cluster_folder,workbook_cluster_traj_folder=workbook_cluster_traj_folder)
			print('--------------------------------------')

	def process_traj_files(self, traj_files_to_process, cluster_information, workbook_cluster_folder='', workbook_cluster_traj_folder=''):

		# process data
		print('###############################################################################')
		for filepath, name in traj_files_to_process:
			print('--------------------------------------')
			excel_name = name+'_processed_data'
			processing_string = 'Processing: '+str(name)
			print(processing_string)
			cluster_images = read(filepath,index=':')
			cluster_images_to_process = [(image, image_number) for image, image_number in zip(cluster_images, range(len(cluster_images)))]
			cluster_information = self.process_cluster_data(cluster_images_to_process)
			analysed_cluster_information = self.analyse_cluster_data(cluster_information)
			self.record_to_excel(cluster_information,analysed_cluster_information,excel_name,combine_clusters_together=True,record_bond_distance=self.record_bond_distance,workbook_cluster_folder=workbook_cluster_folder,workbook_cluster_traj_folder=workbook_cluster_traj_folder)
			print('--------------------------------------')
		print('###############################################################################')


	# ------------------------------------------------------------------------------------------------------------------------------
	# PART 1: OBTAINING DIRECTORY PATH OF CLUSTERS TO PROCESS

	def get_paths_of_all_clusters_to_process(self):
		"""
		This method is designed to obtain all the directories of all the xyz and traj files that this program will run.

		Assume all traj files are make of many images of a cluster from a MD simulation

		"""
		clusters_to_process = []
		traj_files_to_process = []
		for root, dirs, files in os.walk(self.path_to_here):
			if self.record_all_files: # if you want to record all the xyz and traj files in folders, do this
				found_file = False
				for file in files:
					if self.accepted_filetype(file):
						name = file
						if not file.endswith('.traj'):
							cluster = read(root+'/'+file)
							if self.auto_centre:
								self.perform_auto_centre(cluster)
							cluster_to_process = (cluster,name)
							clusters_to_process.append(cluster_to_process)
						else:
							cluster_to_process = (root+'/'+file,name)
							traj_files_to_process.append(cluster_to_process)
						found_file = True
				if found_file:
					# This prevents the search from digging into further directories if at least one xyz or trajectory file is found
					dirs[:] = []
					files[:] = [] 
			else: # if you just want to process one file from each folder, such as processing vasp files and not xyz files, do this
				for file in ['LCGF_look_at.xyz','CONTCAR','OUTCAR']:
					if file in files:
						break
				else:
					continue
				name = root.replace(self.original_path,'')
				cluster = read(root+'/'+file)
				if self.auto_centre:
					self.perform_auto_centre(cluster)
				cluster_to_process = (cluster,name)
				clusters_to_process.append(cluster_to_process)
				# This prevents the search from digging into further directories if one of ['LCGF_look_at.xyz','CONTCAR','OUTCAR'] is found
				dirs[:] = [] 
				files[:] = []

		clusters_to_process.sort(key=lambda x:len(x[0]))
		traj_files_to_process.sort()
		return clusters_to_process, traj_files_to_process

	def accepted_filetype(self,file):
		"""
		Determine if a folder contains an accepted filetype.
		"""
		if file.endswith('.xyz'):
			return True
		elif file.endswith('.traj'):
			return True
		elif file in ['LCGF_look_at.xyz','CONTCAR','OUTCAR']:
			return True
		return False

	def perform_auto_centre(self,system):
		"""
		This method will centre the clusters in the middle of the cell. Needed if your cluster is wrapped across a cell periodic boundary.
		"""
		def get_middle_of_cell(system):
			cell = system.get_cell()
			vec1, vec2, vec3 = cell
			corners = []
			for a,b,c in itertools.product((0,1),(0,1),(0,1)):
				corner_position = a*vec1 + b*vec2 + c*vec3
				corners.append(corner_position)
			centre_of_cell = np.array([sum(corners_axis) for corners_axis in zip(*corners)])/len(corners)
			return centre_of_cell
		add_to_each_position = get_middle_of_cell(system) - system.get_center_of_mass()
		system.set_positions(system.get_positions() + add_to_each_position)
		system.wrap()
		system.center()

	# ------------------------------------------------------------------------------------------------------------------------------
	# PART 2: PROCESS INFORMATION ABOUT EACH CLUSTER

	def process_cluster_data(self,clusters_to_process):
		"""
		This method points to another method for processing nearest neighbour information about a cluster
		"""
		cluster_information = self.process_NN_1(clusters_to_process)
		return cluster_information

	def process_NN_1(self, clusters_to_process):
		"""
		This method is one way to process the nearest neighbour data within a cluster. 

		There was probably other ways of doing this but this was the first, hence why this method ends with thr _1 suffix
		"""
		cluster_information = []
		# if you have tqdm, it will give you a progress bar about where the program is up to in processing
		if does_tqdm_exist:
			clusters_data_for_loop = tqdm(clusters_to_process)
		else:
			clusters_data_for_loop = clusters_to_process

		# For each cluster to be processed, 
		# 1. Get the nearest neighbours about atoms in cluster
		# 2. Obtain the number of atoms that have XX neighbours. 
		for cluster, name in clusters_data_for_loop:
			bond_radii = [self.r_cut/2.0]*len(cluster)
			atoms_neighbours_in_cluster = self.get_no_of_neighbours_for_atoms_in_cluster(cluster, bond_radii)
			all_number_of_neighbours, element_number_of_neighbours = self.get_information_about_number_of_atoms_with_so_many_nn(atoms_neighbours_in_cluster, cluster)
			cluster_information.append((name, element_number_of_neighbours, all_number_of_neighbours, cluster))
		return cluster_information

	def get_no_of_neighbours_for_atoms_in_cluster(self, cluster, bond_radii):
		"""
		This method will create a dictionary that indicates all the neighbours between atoms in a cluster
		"""
		atoms_neighbours_in_cluster = {atom_index: [] for atom_index in range(len(cluster))}
		for atom_1_index in range(len(cluster)):
			atom_1_radius = bond_radii[atom_1_index]
			for atom_2_index in range(atom_1_index+1, len(cluster)):
				atom_2_radius = bond_radii[atom_2_index]
				max_bonding_distance = atom_1_radius + atom_2_radius
				length = cluster.get_distance(atom_1_index,atom_2_index)
				# if two atoms are within the accessable distance to be considered neighbours in a cluster, then they are neighbours
				if length <= max_bonding_distance:
					atoms_neighbours_in_cluster[atom_1_index].append(atom_2_index)
					atoms_neighbours_in_cluster[atom_2_index].append(atom_1_index)
		return atoms_neighbours_in_cluster

	def get_information_about_number_of_atoms_with_so_many_nn(self, atoms_neighbours_in_cluster, cluster):
		"""
		This method will create two dictionaries that indicate the number of atoms that have XX nearest neighbours.

		Will also do this for each element type in the cluster
		"""
		all_number_of_neighbours = {}
		element_number_of_neighbours = {}
		for index in range(len(atoms_neighbours_in_cluster)):
			indices = atoms_neighbours_in_cluster[index]
			number_of_neighbours = len(indices)
			all_number_of_neighbours.setdefault(number_of_neighbours,[]).append(index) # This will include atom index in the column for XX number of nearest neighbours. 
			if cluster[index].symbol == self.focus_plot_with_respect_to_element:
				# If you want to plot graphs focussing on a particular element type, this will include atom index in the column for XX number of nearest neighbours if it is the selected type of element 
				element_number_of_neighbours.setdefault(number_of_neighbours,[]).append(index)
		return all_number_of_neighbours, element_number_of_neighbours

	# ------------------------------------------------------------------------------------------------------------------------------
	# PART 3: FURTHER ANALYSE INFORMATION ABOUT EACH CLUSTER
	# This uses the specifications of the number of neighbours that a bulk, face, edge, and corner atom have in a cluster, as specified by Dr. Caitlin Casey-Stevens. 

	def analyse_cluster_data(self,cluster_information):
		"""
		This method will record which atoms containing XX number of neighbours are assigned as either bulk, face, edge, or corners. 

		This is done for all atoms, and for elements of a specific element. 
		"""
		analysed_cluster_information = []
		# if you have tqdm, this will give you a progress bar for the analysis process
		if does_tqdm_exist:
			cluster_information_for_loop = tqdm(cluster_information)
		else:
			cluster_information_for_loop = cluster_information

		for name, element_number_of_neighbours, all_number_of_neighbours, cluster in cluster_information_for_loop:
			# ---------------------------------------------------------------------
			# Analyse atoms of the selected element type
			analysed_element_number_of_neighbours = {}
			for number_of_neighbours, indices in element_number_of_neighbours.items():
				element_type_name = self.types_of_NNs[self.get_nn_type(number_of_neighbours)]
				self.add_to_dictionary_list(analysed_element_number_of_neighbours,element_type_name,indices)
			# ---------------------------------------------------------------------
			# Analuse all atoms of all element types
			analysed_all_number_of_neighbours = {}
			for number_of_neighbours, indices in all_number_of_neighbours.items():
				element_type_name = self.types_of_NNs[self.get_nn_type(number_of_neighbours)]
				self.add_to_dictionary_list(analysed_all_number_of_neighbours,element_type_name,indices)
			# ---------------------------------------------------------------------
			analysed_cluster_information.append((name, analysed_element_number_of_neighbours, analysed_all_number_of_neighbours, cluster))
		return analysed_cluster_information

	def get_nn_type(self,number_of_neighbours):
		"""
		This will sort the atoms into the appropriate nearest neighbour type:

		0: bulk
		1: face
		2: edge
		3: corner

		"""
		if number_of_neighbours >= 12:
			nn_type = 0
		elif 9 <= number_of_neighbours <= 11:
			nn_type = 1
		elif 7 <= number_of_neighbours <= 8:
			nn_type = 2
		elif number_of_neighbours <= 6:
			nn_type = 3
		else:
			exit('Huh?')
		return nn_type

	def add_to_dictionary_list(self,analysed_element_number_of_neighbours,element_type_name,indices):
		'''
			Add a key to a dictionary
		'''
		analysed_element_number_of_neighbours[element_type_name] = analysed_element_number_of_neighbours.get(element_type_name,[]) + indices


	# ------------------------------------------------------------------------------------------------------------------------------
	# PART 4: SAVE DATA AS EXCEL SPREADSHEETS

	def record_to_excel(self, cluster_information, analysed_cluster_information, workbook_name, combine_clusters_together=False, record_bond_distance=False, workbook_cluster_folder='', workbook_cluster_traj_folder=''):
		"""
		Save data to excel spreadsheet
		"""

		if workbook_cluster_folder == '':
			exit('issue')
		if workbook_cluster_traj_folder == '':
			exit('issue')

		# Create excel spreadsheet
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Info'

		# Methods for helping make excel spreadsheet
		def hex_to_rgb(hex_key):
			rgb = []
			for i in (0,2,4):
				decimal = int(hex_key[i:i+2],16)
				rgb.append(decimal)
			return tuple(rgb)

		def give_black_or_white_writing(hex_key):
			rgb = hex_to_rgb(hex_key)
			record = []
			for c in rgb:
				c = c/255.0
				if c <= 0.03928:
					c = c/12.92
				else:
					c = ((c+0.055)/1.055) ** 2.4
				record.append(c)
			r, g, b = record
			L = 0.2126*r + 0.7152*g + 0.0722*b

			if L > ((1.05*0.05)**0.5 - 0.05):
				return '000000'
			else:
				return 'FFFFFF'

		def get_colour_name(name):
			for colour_name in self.colours.keys():
				if colour_name in name:
					return colour_name
			return 'None'

		# Make the names of columns for excel spreadsheet
		other_namings = [[type_of_NN+': element', type_of_NN+': all', type_of_NN+': percent'] for type_of_NN in self.types_of_NNs]
		naming = ['name', str(tuple(self.elements)), 'No of atoms']+list(itertools.chain.from_iterable(other_namings))
		for index in range(len(naming)):
			name = naming[index]
			worksheet.cell(column=index+1, row=1, value=str(name))
			background_colour = self.colours[get_colour_name(name)].replace('#','')
			worksheet.cell(column=index+1, row=1).font = Font(color=give_black_or_white_writing(background_colour))
			worksheet.cell(column=index+1, row=1).fill = PatternFill("solid", fgColor=background_colour)

		# sort the data by clusters molecular name
		analysed_cluster_information.sort(key=lambda x: (len(x[3]),tuple(value for key, value in sorted(Counter(x[3].get_chemical_symbols()).items()))))

		# record the data for each cluster in for each column
		for index_aci in range(len(analysed_cluster_information)):
			cluster_name, analysed_element_number_of_neighbours, analysed_all_number_of_neighbours, cluster = analysed_cluster_information[index_aci]
			worksheet.cell(column=1, row=index_aci+2, value=str(cluster_name))
			worksheet.cell(column=2, row=index_aci+2, value=str(cluster.get_chemical_formula()))
			worksheet.cell(column=3, row=index_aci+2, value=str(len(cluster)))

			for index2 in range(len(self.types_of_NNs)):
				types_of_NN = self.types_of_NNs[index2]
				element_NN = len(analysed_element_number_of_neighbours[types_of_NN]) if (types_of_NN in analysed_element_number_of_neighbours) else 0
				all_NN     = len(analysed_all_number_of_neighbours[types_of_NN]) if (types_of_NN in analysed_all_number_of_neighbours) else 0

				background_colour = self.colours[get_colour_name(types_of_NN)].replace('#','')

				worksheet.cell(column=3*index2+4, row=index_aci+2, value=str(element_NN))
				worksheet.cell(column=3*index2+4, row=index_aci+2).font = Font(color=give_black_or_white_writing(background_colour))
				worksheet.cell(column=3*index2+4, row=index_aci+2).fill = PatternFill("solid", fgColor=background_colour)

				worksheet.cell(column=3*index2+5, row=index_aci+2, value=str(all_NN))
				worksheet.cell(column=3*index2+5, row=index_aci+2).font = Font(color=give_black_or_white_writing(background_colour))
				worksheet.cell(column=3*index2+5, row=index_aci+2).fill = PatternFill("solid", fgColor=background_colour)

				percentage = (float(element_NN)/float(all_NN))*100.0
				worksheet.cell(column=3*index2+6, row=index_aci+2, value=str(percentage))
				worksheet.cell(column=3*index2+6, row=index_aci+2).font = Font(color=give_black_or_white_writing(background_colour))
				worksheet.cell(column=3*index2+6, row=index_aci+2).fill = PatternFill("solid", fgColor=background_colour)

		###############################################################
		# write distances into excel

		# process the bond length between each atom in the cluster. 
		if record_bond_distance:
			print('Writing bond distance data to excel')
			# if you have tqdm, this will give you a progress bar for the analysis process
			if does_tqdm_exist:
				range_for_loop = trange(len(analysed_cluster_information))
			else:
				range_for_loop = range(len(analysed_cluster_information))

			for index_aci in range_for_loop:
				cluster_path, analysed_element_number_of_neighbours, analysed_all_number_of_neighbours, cluster = analysed_cluster_information[index_aci]
				if self.record_all_files:
					cluster_name = cluster_path.replace('.','_') ####### to do
				else:
					cluster_name = str(len(cluster))+'_'+str(cluster.get_chemical_formula())+'_'+cluster_path.split('/')[-1]
				if does_tqdm_exist:
					range_for_loop.set_description("Processing "+str(cluster_name))
				else:
					print(str(index_aci+1)+' out of '+str(len(analysed_cluster_information))+' ('+str(cluster_name)+')')
				# make a new spreadsheet.
				worksheet = workbook.create_sheet()
				worksheet.title = str(cluster_name)

				# Write in the cell the name for data
				worksheet.cell(row=1, column=1).value = 'r_cut = '
				worksheet.cell(row=1, column=2).value = str(self.r_cut)
				worksheet.cell(row=3, column=3).value = 'atom index'
				worksheet.cell(row=3, column=2).value = worksheet.cell(row=2, column=3).value = 'no of neighbours'
				worksheet.cell(row=3, column=1).value = worksheet.cell(row=1, column=3).value = 'element'
				symbols = cluster.get_chemical_symbols()
				# Write in the row and column for the index of each atom
				for index1 in range(len(cluster)):
					worksheet.cell(row=4+index1, column=1).value = worksheet.cell(row=1, column=4+index1).value = str(symbols[index1])
					worksheet.cell(row=4+index1, column=3).value = worksheet.cell(row=3, column=4+index1).value = str(index1)
				no_of_nearest_neighbours_per_atom = {index: 0 for index in range(len(cluster))}

				# write the distances between atoms in cluster
				cluster_positions = cluster.get_positions()
				for index1 in range(len(cluster)):
					for index2 in range(index1+1,len(cluster)):
						distance = get_distance(cluster_positions[index1],cluster_positions[index2])
						worksheet.cell(row=4+index1, column=4+index2).value = worksheet.cell(row=4+index2, column=4+index1).value = str(round(distance,3))
						# Make a note if the distance between atoms is within r_cut or not. 
						if distance <= self.r_cut:
							no_of_nearest_neighbours_per_atom[index1] += 1
							no_of_nearest_neighbours_per_atom[index2] += 1
							worksheet.cell(row=4+index1, column=4+index2).font = Font(color=give_black_or_white_writing('90EE90'))
							worksheet.cell(row=4+index1, column=4+index2).fill = worksheet.cell(row=4+index2, column=4+index1).fill = PatternFill("solid", fgColor='90EE90')
						else:
							worksheet.cell(row=4+index1, column=4+index2).font = Font(color=give_black_or_white_writing('FFC0CB'))
							worksheet.cell(row=4+index1, column=4+index2).fill = worksheet.cell(row=4+index2, column=4+index1).fill = PatternFill("solid", fgColor='FFC0CB')
					worksheet.cell(row=4+index1, column=2).value = worksheet.cell(row=2, column=4+index1).value = str(no_of_nearest_neighbours_per_atom[index1])

				# Write the number of neighbours on each atom in the traj file, and write the traj file.
				nn_atoms = [nn for index, nn in sorted(no_of_nearest_neighbours_per_atom.items())]
				cluster.set_tags(nn_atoms)
				xyz_file_name = cluster_name+'_no_of_neighbours'
				if combine_clusters_together:
					images_of_cluster_from_MD_sim_no_of_neighbours.append(cluster.copy())
				else:
					write(workbook_cluster_folder+'/'+xyz_file_name+'.xyz',cluster)

				# The types of vertices on each atom
				type_of_nn_atoms = [self.get_nn_type(nn) for index, nn in sorted(no_of_nearest_neighbours_per_atom.items())]
				cluster.set_tags(type_of_nn_atoms)
				xyz_file_name = cluster_name+'_nn_type'
				if combine_clusters_together:
					images_of_cluster_from_MD_sim_nn_type.append(cluster.copy())
				else:
					write(workbook_cluster_folder+'/'+xyz_file_name+'.xyz',cluster)

				# Save the filename of clusters that have been processed.
				if not combine_clusters_together:
					with open(workbook_cluster_folder+'/filenames_to_paths.txt','a') as xyz_names_to_path:
						xyz_names_to_path.write(str(cluster_name)+': '+str(cluster_path)+'\n')

		# write traj files of no_of_neighbours and nn_type 
		if combine_clusters_together:
			images_of_cluster_from_MD_sim_no_of_neighbours = []
			images_of_cluster_from_MD_sim_nn_type = []
			for name, analysed_element_number_of_neighbours, analysed_all_number_of_neighbours, cluster in cluster_information:
				# get no_of_neighbours tags in cluster traj
				cluster_no_of_neighbours = cluster.copy()
				no_of_neighbours_indices = [None for index in range(len(cluster_no_of_neighbours))]
				for no_of_neighbours, indices_with_no_of_neighbours in analysed_all_number_of_neighbours.items():
					for index_with_no_of_neighbours in indices_with_no_of_neighbours:
						no_of_neighbours_indices[index_with_no_of_neighbours] = no_of_neighbours
				cluster_no_of_neighbours.set_tags(no_of_neighbours_indices)
				images_of_cluster_from_MD_sim_no_of_neighbours.append(cluster_no_of_neighbours.copy())

				# get type_of_nn tags in cluster traj
				cluster_nn_type = cluster.copy()
				nn_type_indices = [self.get_nn_type(nn) for nn in no_of_neighbours_indices]
				cluster_nn_type.set_tags(nn_type_indices)
				images_of_cluster_from_MD_sim_nn_type.append(cluster_nn_type.copy())

			write(workbook_cluster_traj_folder+'/'+workbook_name.replace('.','_')+'_no_of_neighbours.traj', images_of_cluster_from_MD_sim_no_of_neighbours)
			write(workbook_cluster_traj_folder+'/'+workbook_name.replace('.','_')+'_nn_type.traj',          images_of_cluster_from_MD_sim_nn_type)

		# Save the file
		print('Saving excel file')
		if combine_clusters_together:
			workbook.save(workbook_cluster_traj_folder+'/'+workbook_name.replace('.','_')+".xlsx")
		else:
			workbook.save(workbook_name+".xlsx")
		print('Done saving excel file')

	# ------------------------------------------------------------------------------------------------------------------------------




